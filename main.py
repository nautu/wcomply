import io
import logging
import os
import re
import threading
import time
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Optional

import openpyxl
from bson import ObjectId
from dotenv import load_dotenv
from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from jose import JWTError, jwt
from passlib.context import CryptContext
from pymongo import ASCENDING, DESCENDING, MongoClient
from starlette.middleware.base import BaseHTTPMiddleware

load_dotenv()

# ── Security logger ───────────────────────────────────────────

def _setup_security_logger() -> logging.Logger:
    logger = logging.getLogger("vulntrack.security")
    logger.setLevel(logging.INFO)
    log_dir = "/var/log/vulntrack"
    try:
        os.makedirs(log_dir, exist_ok=True)
        fh = logging.FileHandler(f"{log_dir}/security.log")
        fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        logger.addHandler(fh)
    except PermissionError:
        pass  # fallback to stdout only
    sh = logging.StreamHandler()
    sh.setFormatter(logging.Formatter("%(asctime)s SECURITY %(message)s"))
    logger.addHandler(sh)
    return logger

security_log = _setup_security_logger()

_MONGO_OP_RE = re.compile(r'\$[a-zA-Z]')

# ── Config ────────────────────────────────────────────────────
MONGO_URL     = os.getenv("MONGO_URL", "mongodb://localhost:27017")
MONGO_DB_NAME = os.getenv("MONGO_DB_NAME", "vulntrack")
SECRET_KEY    = os.getenv("SECRET_KEY", "change_me_in_production")
ALGORITHM     = "HS256"
TOKEN_HOURS   = 8

_mongo = MongoClient(MONGO_URL, serverSelectionTimeoutMS=5000)
_db    = _mongo[MONGO_DB_NAME]

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto", bcrypt__truncate_error=False)


def get_db():
    return _db


# ── Import job (statut persisté en base, partagé entre workers) ─
_JOB_ID = "advisory_import"


def _job_get() -> dict:
    doc = _db.jobs.find_one({"_id": _JOB_ID})
    return doc or {"status": "idle", "msg": "", "detail": ""}


def _job_set(**fields):
    _db.jobs.update_one({"_id": _JOB_ID}, {"$set": fields}, upsert=True)


# ── App ───────────────────────────────────────────────────────
MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB

app = FastAPI(title="VulnTrack")
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")


class UploadSizeLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        cl = request.headers.get("content-length")
        if cl and int(cl) > MAX_UPLOAD_BYTES:
            from fastapi.responses import JSONResponse
            return JSONResponse(
                {"detail": f"Fichier trop volumineux — limite : {MAX_UPLOAD_BYTES // 1024 // 1024} MB"},
                status_code=413,
            )
        return await call_next(request)


app.add_middleware(UploadSizeLimitMiddleware)


# ── Security helpers ──────────────────────────────────────────

XLSX_MAGIC = b'PK\x03\x04'
_SUSPICIOUS_NAME_RE = re.compile(r'[/\\<>:"|?*\x00]|\.\.')

_UPLOAD_WINDOWS: dict = defaultdict(list)
_API_WINDOWS: dict = defaultdict(list)
_rl_lock = threading.Lock()

UPLOAD_LIMIT = 10
UPLOAD_WINDOW = 3600  # 1 hour

API_LIMIT = 100
API_WINDOW = 60  # 1 minute

_UPLOAD_PATHS = {"/advisory/upload", "/frun/upload"}


def _get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _check_rate_limit(store: dict, ip: str, limit: int, window: float) -> bool:
    now = time.time()
    with _rl_lock:
        store[ip] = [t for t in store[ip] if now - t < window]
        if len(store[ip]) >= limit:
            return False
        store[ip].append(now)
        return True


def _validate_xlsx(filename: str, content: bytes, ip: str) -> None:
    if not filename.lower().endswith(".xlsx"):
        security_log.warning(f"Upload rejected non-xlsx ip={ip} filename={filename!r}")
        raise HTTPException(400, "Seuls les fichiers .xlsx sont acceptés")
    if _SUSPICIOUS_NAME_RE.search(filename) or len(filename) > 255:
        security_log.warning(f"Upload rejected suspicious filename ip={ip} filename={filename!r}")
        raise HTTPException(400, "Nom de fichier non autorisé")
    if not content.startswith(XLSX_MAGIC):
        security_log.warning(f"Upload rejected invalid magic bytes ip={ip} filename={filename!r}")
        raise HTTPException(400, "Format de fichier invalide (non reconnu comme xlsx)")


def _sanitize_text(value: str, field: str = "") -> str:
    """Reject MongoDB operator injection; strip whitespace."""
    if not isinstance(value, str):
        return value
    if _MONGO_OP_RE.search(value):
        security_log.warning(f"MongoDB operator injection attempt field={field!r} value={value[:100]!r}")
        raise HTTPException(400, "Valeur non autorisée")
    return value.strip()


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        qs = str(request.url.query)
        path = request.url.path
        if _MONGO_OP_RE.search(qs) or ".." in path:
            ip = _get_client_ip(request)
            security_log.warning(
                f"Suspicious request ip={ip} method={request.method} "
                f"path={path!r} qs={qs[:200]!r}"
            )
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src https://fonts.gstatic.com"
        )
        return response


class RateLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        from fastapi.responses import JSONResponse
        path = request.url.path
        if path.startswith("/static/"):
            return await call_next(request)
        ip = _get_client_ip(request)
        if request.method == "POST" and path in _UPLOAD_PATHS:
            if not _check_rate_limit(_UPLOAD_WINDOWS, ip, UPLOAD_LIMIT, UPLOAD_WINDOW):
                security_log.warning(f"Upload rate limit exceeded ip={ip}")
                return JSONResponse(
                    {"detail": "Trop de tentatives d'upload. Réessayez dans une heure."},
                    status_code=429,
                )
        elif not _check_rate_limit(_API_WINDOWS, ip, API_LIMIT, API_WINDOW):
            security_log.warning(f"API rate limit exceeded ip={ip} path={path!r}")
            return JSONResponse(
                {"detail": "Trop de requêtes. Réessayez dans une minute."},
                status_code=429,
            )
        return await call_next(request)


app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(RateLimitMiddleware)


# ── Auth utilities ────────────────────────────────────────────

def _hash_pw(pw: str) -> str:
    return pwd_ctx.hash(pw)


def _verify_pw(pw: str, hashed: str) -> bool:
    return pwd_ctx.verify(pw, hashed)


def _create_token(username: str, role: str) -> str:
    exp = datetime.now(timezone.utc) + timedelta(hours=TOKEN_HOURS)
    return jwt.encode({"sub": username, "role": role, "exp": exp},
                      SECRET_KEY, algorithm=ALGORITHM)


def _decode_token(token: str) -> dict:
    return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])


def _init_admin():
    if _db.users.count_documents({}) == 0:
        pw = "Admin2026"
        _db.users.insert_one({
            "username":      "admin",
            "password_hash": _hash_pw(pw),
            "role":          "admin",
            "created_at":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
        print("")
        print("[VulnTrack] ★  Premier démarrage — compte admin créé")
        print("[VulnTrack]    username : admin")
        print(f"[VulnTrack]    password : {pw}")
        print("[VulnTrack]    → Changez ce mot de passe via /admin dès que possible.")
        print("")


# ── Auth Middleware — DÉSACTIVÉ (bypass temporaire) ───────────
# Pour réactiver : décommenter le bloc ci-dessous et supprimer
# le dict _BYPASS_USER dans get_current_user / require_admin.
#
# class AuthMiddleware(BaseHTTPMiddleware):
#     async def dispatch(self, request: Request, call_next):
#         path = request.url.path
#         if path == "/login" or path.startswith("/static/"):
#             return await call_next(request)
#
#         token = request.cookies.get("vt_token")
#         if not token:
#             return RedirectResponse("/login", status_code=302)
#
#         try:
#             payload = _decode_token(token)
#             request.state.user = payload
#         except JWTError:
#             resp = RedirectResponse("/login", status_code=302)
#             resp.delete_cookie("vt_token")
#             return resp
#
#         return await call_next(request)
#
# app.add_middleware(AuthMiddleware)


# ── Startup ───────────────────────────────────────────────────

@app.on_event("startup")
def create_indexes():
    try:
        _db.sap_notes.create_index([("reference_note",   ASCENDING)])
        _db.sap_notes.create_index([("advisory_release", ASCENDING)])
        _db.frun_data.create_index([("client",    ASCENDING)])
        _db.frun_data.create_index([("check_ref", ASCENDING)])
        _db.users.create_index([("username", ASCENDING)], unique=True)
        _init_admin()
        # Un job "running" au démarrage est forcément orphelin (service redémarré
        # pendant un import) — on le remet à idle pour débloquer l'UI.
        _db.jobs.update_one(
            {"_id": _JOB_ID, "status": "running"},
            {"$set": {"status": "idle", "msg": "", "detail": ""}},
        )
    except Exception as e:
        print(f"[startup] MongoDB non disponible ({e}). Les index seront créés à la première connexion.")


STATUS_VALUES = [
    "Not Started", "Started", "Implemented",
    "Validated", "To be checked", "Exception",
]


# ── Route helper & dependencies ───────────────────────────────

def tpl(request: Request, name: str, ctx: Optional[dict] = None):
    user = getattr(request.state, "user", {})
    data: dict = {"current_user": user}
    if ctx:
        data.update(ctx)
    return templates.TemplateResponse(request, name, data)


# Bypass temporaire : toutes les routes sont ouvertes sans login.
# Pour réactiver l'auth, remplacer ces deux fonctions par :
#   def get_current_user(request): return getattr(request.state, "user", {})
#   def require_admin(user=Depends(get_current_user)):
#       if user.get("role") != "admin": raise HTTPException(403)
#       return user
_BYPASS_USER = {"sub": "admin", "role": "admin"}


def get_current_user(request: Request) -> dict:
    return _BYPASS_USER


def require_admin(user: dict = Depends(get_current_user)):
    return user


# ── Business logic utilities ──────────────────────────────────

def clean_check_ref(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"\D", "", str(value))
    return int(digits) if digits else None


def calculate_priority(score) -> Optional[str]:
    if score is None:
        return None
    if score >= 9:
        return "P1"
    if score >= 7:
        return "P2"
    return "P3"


def format_advisory_release(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m")
    # Excel serial date (days since 1899-12-30, accounts for the 1900 leap-year bug)
    if isinstance(val, (int, float)) and val > 1:
        try:
            d = date(1899, 12, 30) + timedelta(days=int(val))
            if d.year >= 2000:
                return d.strftime("%Y-%m")
        except Exception:
            pass
    s = str(val).strip()
    if re.match(r"\d{4}-\d{2}", s):
        return s[:7]
    return None


_CVSS_PREFIX_RE = re.compile(r'^\[p[1-3]-cvss\s+[\d.]+\]\s*', re.IGNORECASE)


def strip_cvss_prefix(s: Optional[str]) -> Optional[str]:
    if not s:
        return s
    return _CVSS_PREFIX_RE.sub('', s).strip() or s


def apply_dedup(db):
    pipeline = [
        {"$sort": {"version": DESCENDING}},
        {"$group": {
            "_id": "$reference_note",
            "keep_id": {"$first": "$_id"},
            "count":   {"$sum": 1},
        }},
        {"$match": {"count": {"$gt": 1}}},
    ]
    for group in db.sap_notes.aggregate(pipeline):
        db.sap_notes.delete_many({
            "reference_note": group["_id"],
            "_id": {"$ne": group["keep_id"]},
        })


def build_merged_entries(db, client=None, sid=None, reference=None) -> list:
    filt: dict = {}
    if client:
        filt["client"] = {"$regex": re.escape(client), "$options": "i"}
    if sid:
        filt["sid"] = {"$regex": re.escape(sid), "$options": "i"}
    if reference:
        ref_num = clean_check_ref(reference)
        if ref_num:
            filt["check_ref"] = ref_num

    frun_list = list(
        db.frun_data.find(filt).sort([("client", ASCENDING), ("sid", ASCENDING)])
    )
    if not frun_list:
        return []

    check_refs = list({e["check_ref"] for e in frun_list if e.get("check_ref")})
    advisory_map: dict = {}
    if check_refs:
        for note in db.sap_notes.find({"reference_note": {"$in": check_refs}}):
            advisory_map[note["reference_note"]] = note

    results = []
    for frun in frun_list:
        adv   = advisory_map.get(frun.get("check_ref"))
        score = adv.get("cvss_v3_base_score") if adv else None
        results.append({
            "id":             str(frun["_id"]),
            "client":         frun.get("client"),
            "priority":       calculate_priority(score),
            "sid":            frun.get("sid"),
            "status":         frun.get("status", "Not Started"),
            "comment":        frun.get("comment", ""),
            "status_history": frun.get("status_history", []),
            "check_description":                  strip_cvss_prefix(frun.get("check_description")),
            "policy":                             frun.get("policy"),
            "correction_type":                    adv.get("correction_type")                    if adv else None,
            "reference":                          adv.get("reference_note")                     if adv else frun.get("check_ref"),
            "cve":                                adv.get("cve")                                if adv else None,
            "type":                               (adv.get("type") or adv.get("category"))      if adv else None,
            "recommended_implementation_process": adv.get("recommended_implementation_process") if adv else None,
            "downtime_required":                  adv.get("downtime_required")                  if adv else None,
            "solution_short":                     adv.get("solution_short")                     if adv else None,
            "workaround":                         adv.get("workaround")                         if adv else None,
            "landscape":          frun.get("landscape"),
            "configuration_item": frun.get("configuration_item"),
            "value":              frun.get("value"),
            "rule":               frun.get("rule"),
            "valid_since_utc":    frun.get("valid_since_utc"),
            "upload_timestamp":   frun.get("imported_at"),
            "cvss_score":         score,
        })
    return results


def _next_release(current: Optional[str]) -> Optional[str]:
    if not current or current == "—":
        return None
    try:
        year, month = map(int, current.split("-"))
        year, month = (year + 1, 1) if month == 12 else (year, month + 1)
        return f"{year}-{month:02d}"
    except Exception:
        return None


def _client_summary(entries: list, client: str) -> dict:
    total = len(entries)
    p1 = sum(1 for e in entries if e["priority"] == "P1")
    p2 = sum(1 for e in entries if e["priority"] == "P2")
    p3 = sum(1 for e in entries if e["priority"] == "P3")

    by_status: dict[str, int] = {}
    for e in entries:
        by_status[e["status"]] = by_status.get(e["status"], 0) + 1

    done        = sum(1 for e in entries if e["status"] in ("Validated", "Implemented"))
    in_progress = sum(1 for e in entries if e["status"] == "Started")
    not_started = by_status.get("Not Started", 0)
    progress_pct = round(done / total * 100) if total else 0

    critical = [e for e in entries if e["priority"] == "P1"
                and e["status"] in ("Not Started", "Started")]

    sid_map: dict = {}
    for e in entries:
        sid = e["sid"] or "—"
        if sid not in sid_map:
            sid_map[sid] = {"sid": sid, "total": 0, "p1": 0, "p2": 0, "p3": 0,
                            "done": 0, "by_status": {}}
        d = sid_map[sid]
        d["total"] += 1
        if e["priority"] == "P1":   d["p1"] += 1
        elif e["priority"] == "P2": d["p2"] += 1
        elif e["priority"] == "P3": d["p3"] += 1
        if e["status"] in ("Validated", "Implemented"): d["done"] += 1
        d["by_status"][e["status"]] = d["by_status"].get(e["status"], 0) + 1

    for d in sid_map.values():
        d["progress_pct"] = round(d["done"] / d["total"] * 100) if d["total"] else 0

    return {
        "client": client, "total": total,
        "p1": p1, "p2": p2, "p3": p3,
        "no_score": total - p1 - p2 - p3,
        "by_status": by_status,
        "done": done, "in_progress": in_progress, "not_started": not_started,
        "progress_pct": progress_pct,
        "critical_count": len(critical),
        "sid_breakdown": sorted(sid_map.values(), key=lambda x: (-x["p1"], -x["total"])),
        "sids": [d["sid"] for d in sorted(sid_map.values(), key=lambda x: (-x["p1"], -x["total"]))],
    }


# ── Routes : Authentification — DÉSACTIVÉES (bypass temporaire) ─
# Pour réactiver, décommenter ce bloc et rétablir le middleware + dépendances.
#
# @app.get("/login", response_class=HTMLResponse)
# def login_page(request: Request, error: str = ""):
#     token = request.cookies.get("vt_token")
#     if token:
#         try:
#             _decode_token(token)
#             return RedirectResponse("/", status_code=302)
#         except JWTError:
#             pass
#     return templates.TemplateResponse(request, "login.html",
#                                       {"error": error, "current_user": {}})
#
# @app.post("/login")
# async def login(request: Request, db=Depends(get_db)):
#     form     = await request.form()
#     username = str(form.get("username", "")).strip()
#     password = str(form.get("password", ""))
#     user_doc = db.users.find_one({"username": username})
#     if not user_doc or not _verify_pw(password, user_doc["password_hash"]):
#         return templates.TemplateResponse(request, "login.html", {
#             "error": "Identifiants incorrects", "current_user": {}}, status_code=401)
#     token = _create_token(user_doc["username"], user_doc["role"])
#     resp  = RedirectResponse("/", status_code=303)
#     resp.set_cookie("vt_token", token, httponly=True, samesite="lax",
#                     max_age=TOKEN_HOURS * 3600)
#     return resp
#
# @app.post("/logout")
# def logout():
#     resp = RedirectResponse("/login", status_code=303)
#     resp.delete_cookie("vt_token")
#     return resp


# ── Routes : Admin ────────────────────────────────────────────

@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request, msg: str = "", db=Depends(get_db),
               _: dict = Depends(require_admin)):
    users = list(db.users.find({}, {"password_hash": 0})
                 .sort("username", ASCENDING))
    return tpl(request, "admin.html", {"users": users, "msg": msg})


@app.post("/admin/users")
async def create_user(request: Request, db=Depends(get_db),
                      _: dict = Depends(require_admin)):
    form     = await request.form()
    username = str(form.get("username", "")).strip().lower()
    password = str(form.get("password", ""))
    role     = str(form.get("role", "consultant"))

    if not username or not password:
        return RedirectResponse("/admin?msg=Username+et+mot+de+passe+requis",
                                status_code=303)
    if role not in ("admin", "consultant"):
        role = "consultant"
    if db.users.find_one({"username": username}):
        return RedirectResponse(f"/admin?msg=L'utilisateur+'{username}'+existe+déjà",
                                status_code=303)

    db.users.insert_one({
        "username":      username,
        "password_hash": _hash_pw(password),
        "role":          role,
        "created_at":    datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    return RedirectResponse(f"/admin?msg=Compte+'{username}'+créé+avec+succès",
                            status_code=303)


@app.post("/admin/users/{username}/delete")
def delete_user(username: str, db=Depends(get_db),
                admin: dict = Depends(require_admin)):
    if username == admin.get("sub"):
        return RedirectResponse(
            "/admin?msg=Impossible+de+supprimer+votre+propre+compte",
            status_code=303)
    db.users.delete_one({"username": username})
    return RedirectResponse(f"/admin?msg=Compte+'{username}'+supprimé",
                            status_code=303)


# ── Routes : Paramètres ───────────────────────────────────────

@app.post("/settings/reset-db")
def reset_db(db=Depends(get_db), _: dict = Depends(require_admin)):
    db.frun_data.drop()
    db.drop_collection("merged_view")
    try:
        db.frun_data.create_index([("client",    ASCENDING)])
        db.frun_data.create_index([("check_ref", ASCENDING)])
    except Exception:
        pass
    return RedirectResponse("/", status_code=303)


# ── Routes : Dashboard ────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def index(request: Request, db=Depends(get_db)):
    advisory_count = db.sap_notes.count_documents({})
    frun_count     = db.frun_data.count_documents({})
    clients_list   = db.frun_data.distinct("client")
    doc = db.sap_notes.find_one({}, sort=[("advisory_release", DESCENDING)])
    max_release = doc["advisory_release"] if doc else "—"

    all_entries = build_merged_entries(db) if frun_count > 0 else []
    total = len(all_entries)
    p1    = sum(1 for e in all_entries if e["priority"] == "P1")
    p2    = sum(1 for e in all_entries if e["priority"] == "P2")
    p3    = sum(1 for e in all_entries if e["priority"] == "P3")
    done  = sum(1 for e in all_entries if e["status"] in ("Validated", "Implemented"))
    progress_pct = round(done / total * 100) if total else 0

    by_status: dict[str, int] = {}
    for e in all_entries:
        by_status[e["status"]] = by_status.get(e["status"], 0) + 1

    client_p1: dict[str, int] = {}
    for e in all_entries:
        if e["priority"] == "P1" and e.get("client"):
            client_p1[e["client"]] = client_p1.get(e["client"], 0) + 1
    top_client = max(client_p1.items(), key=lambda x: x[1]) if client_p1 else None

    return tpl(request, "index.html", {
        "advisory_count":  advisory_count,
        "frun_count":      frun_count,
        "clients":         len(clients_list),
        "max_release":     max_release,
        "dash_total":      total,
        "dash_p1":         p1,
        "dash_p2":         p2,
        "dash_p3":         p3,
        "dash_done":       done,
        "dash_progress":   progress_pct,
        "dash_by_status":  by_status,
        "dash_top_client": top_client,
        "status_order":    STATUS_VALUES,
    })


# ── Routes : Advisory ─────────────────────────────────────────

@app.get("/advisory", response_class=HTMLResponse)
def advisory_page(request: Request, msg: str = "", db=Depends(get_db)):
    count = db.sap_notes.count_documents({})
    doc   = db.sap_notes.find_one({}, sort=[("advisory_release", DESCENDING)])
    mr    = doc["advisory_release"] if doc else "—"
    return tpl(request, "advisory.html", {
        "count":        count,
        "max_release":  mr,
        "next_release": _next_release(mr if mr != "—" else None),
        "msg":          msg,
    })


def _flush_batch(db, batch: dict) -> tuple[int, int, list]:
    """Écrit un batch de docs Advisory en base.

    - Vérifie chaque reference_note : insert si absent, replace si version plus haute.
    - Les nouveaux documents sont groupés en insert_many (un seul aller-retour réseau).
    - Retourne (added, updated, inserted_ids) pour le suivi du rollback.
    """
    to_insert: list = []
    inserted_ids: list = []
    updated = 0

    for ref, data in batch.items():
        existing = db.sap_notes.find_one({"reference_note": ref}, {"_id": 1, "version": 1})
        if existing:
            if data["version"] > existing.get("version", 0):
                db.sap_notes.replace_one({"reference_note": ref}, data)
                updated += 1
        else:
            to_insert.append(data)

    if to_insert:
        result = db.sap_notes.insert_many(to_insert, ordered=False)
        inserted_ids = list(result.inserted_ids)

    return len(to_insert), updated, inserted_ids


def _process_advisory_import(content: bytes) -> None:
    """Traitement Advisory en arrière-plan (BackgroundTask).

    Optimisations mémoire :
    - read_only=True  → openpyxl streaming XML, pas de chargement en RAM du classeur entier
    - Batch de 100    → accumule 100 entrées uniques avant d'écrire, libère le buffer
    - insert_many     → un seul aller-retour MongoDB par batch

    Rollback manuel : les _id insérés sont trackés ; en cas d'exception,
    delete_many les supprime. (Les transactions multi-documents MongoDB
    nécessitent un replica set — incompatible avec un déploiement standalone.)
    """
    db = _db
    wb = None
    inserted_ids: list = []  # pour le rollback

    _job_set(status="running", detail="Lecture du fichier Excel…", msg="")

    try:
        # read_only=True : parser XML streaming — n'alloue pas tout le classeur en RAM
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active

        _job_set(detail="Détection du format…")
        doc = db.sap_notes.find_one({}, sort=[("advisory_release", DESCENDING)])
        max_release = doc["advisory_release"] if doc else None

        # Lecture de la ligne d'en-tête via l'itérateur (compatible read_only)
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            _job_set(status="error", msg="Fichier vide ou illisible", detail="")
            return

        headers = [str(v).strip() if v is not None else "" for v in header_row]
        col = {h: i for i, h in enumerate(headers)}

        if "Nombre" in col:
            fmt = "fr"
        elif "Number" in col or "Reference Note" in col:
            fmt = "en"
        else:
            _job_set(status="error",
                     msg="Format non reconnu (colonne 'Number' / 'Nombre' / 'Reference Note' attendue)",
                     detail="")
            return

        fmt_label = "SAP for Me" if fmt == "fr" else "anglais"

        def _s(v):
            return str(v).strip() if v is not None else None

        def _get(row, *names):
            for name in names:
                idx = col.get(name)
                if idx is not None:
                    return row[idx] if idx < len(row) else None
            for name in names:
                pfx = name.lower()
                for h, idx in col.items():
                    if h.lower().startswith(pfx) and idx < len(row):
                        return row[idx]
            return None

        def _cvss(raw):
            if raw is None or raw == "-":
                return None
            if isinstance(raw, str):
                try:
                    return float(raw.replace(",", "."))
                except Exception:
                    return None
            try:
                return float(raw)
            except Exception:
                return None

        BATCH_SIZE = 100
        batch: dict[int, dict] = {}   # reference_note → data (dédupliqué dans le batch)
        added = updated = skipped = row_num = 0

        _job_set(detail=f"Traitement des lignes (format {fmt_label})…")

        for row in rows_iter:
            row_num += 1

            if fmt == "en":
                advisory_release = format_advisory_release(_get(row, "Advisory Release"))
            else:
                advisory_release = format_advisory_release(_get(row, "Date/Heure de validation"))

            if not advisory_release:
                continue
            if max_release and advisory_release <= max_release:
                skipped += 1
                continue

            raw_ref = _get(row, "Number", "Reference Note") if fmt == "en" else _get(row, "Nombre")
            if raw_ref is None:
                continue
            try:
                reference_note = int(raw_ref)
            except (ValueError, TypeError):
                continue

            if fmt == "en":
                try:
                    version = int(_get(row, "Version") or 0)
                except (ValueError, TypeError):
                    version = 0

                data = {
                    "advisory_release":    advisory_release,
                    "sap_component":       _s(_get(row, "SAP Component")),
                    "reference_note":      reference_note,
                    "version":             version,
                    "title":               _s(_get(row, "Title")),
                    "category":            _s(_get(row, "Category")),
                    "priority_sap":        _s(_get(row, "Priority")),
                    "cvss_v3_base_score":  _cvss(_get(row, "CVSSv3 Base Score*", "CVSSv3 Base Score")),
                    "correction_type":     _s(_get(row, "Correction type")),
                    "recommended_implementation_process": _s(_get(row, "Recommended implementation process")),
                    "downtime_required":   _s(_get(row, "Downtime required**", "Downtime required")),
                    "cve":                 _s(_get(row, "CVE")),
                    "type":                _s(_get(row, "Type of vulnerability addressed")),
                    "solution_short":      _s(_get(row, "Solution [short]")),
                    "workaround":          _s(_get(row, "Workaround")),
                }
                if reference_note not in batch or version > batch[reference_note]["version"]:
                    batch[reference_note] = data
            else:
                categorie = _s(_get(row, "Catégorie"))
                data = {
                    "advisory_release":    advisory_release,
                    "sap_component":       _s(_get(row, "Composant SAP")),
                    "reference_note":      reference_note,
                    "version":             0,
                    "title":               _s(_get(row, "Titre")),
                    "category":            categorie,
                    "priority_sap":        _s(_get(row, "Priorité")),
                    "cvss_v3_base_score":  _cvss(_get(row, "Score CVSS")),
                    "correction_type":     categorie,
                    "recommended_implementation_process": None,
                    "downtime_required":   None,
                    "cve":                 None,
                    "type":                None,
                    "solution_short":      None,
                    "workaround":          None,
                }
                if reference_note not in batch:
                    batch[reference_note] = data

            # Flush dès que le batch atteint BATCH_SIZE entrées uniques
            if len(batch) >= BATCH_SIZE:
                a, u, ids = _flush_batch(db, batch)
                added += a; updated += u; inserted_ids.extend(ids)
                batch.clear()
                _job_set(detail=f"{row_num} lignes lues — {added} ajoutées, {updated} mises à jour…")

        # Flush du reste
        if batch:
            a, u, ids = _flush_batch(db, batch)
            added += a; updated += u; inserted_ids.extend(ids)

        _job_set(detail="Déduplication finale…")
        apply_dedup(db)

        msg = f"{added} notes ajoutées, {updated} mises à jour, {skipped} ignorées (format {fmt_label})"
        _job_set(status="done", msg=msg, detail="")

    except Exception as e:
        # Rollback : supprime tous les documents nouvellement insérés
        if inserted_ids:
            try:
                db.sap_notes.delete_many({"_id": {"$in": inserted_ids}})
            except Exception:
                pass
        _job_set(status="error",
                 msg=f"Erreur — rollback de {len(inserted_ids)} document(s) inséré(s) : {e}",
                 detail="")
    finally:
        if wb is not None:
            wb.close()


@app.post("/advisory/upload")
async def upload_advisory(request: Request, background_tasks: BackgroundTasks,
                          file: UploadFile = File(...)):
    ip = _get_client_ip(request)
    job = _job_get()
    if job.get("status") == "running":
        return RedirectResponse("/advisory?importing=1", status_code=303)

    content = await file.read()
    try:
        _validate_xlsx(file.filename or "", content, ip)
    except HTTPException as exc:
        security_log.warning(f"Advisory upload rejected ip={ip} reason={exc.detail!r}")
        return RedirectResponse(f"/advisory?msg={exc.detail}", status_code=303)

    _job_set(status="running", msg="", detail="Démarrage…")
    background_tasks.add_task(_process_advisory_import, content)
    return RedirectResponse("/advisory?importing=1", status_code=303)


@app.get("/advisory/import-status")
def advisory_import_status():
    job = _job_get()
    return {"status": job.get("status", "idle"),
            "msg":    job.get("msg", ""),
            "detail": job.get("detail", "")}


# ── Routes : FRun ─────────────────────────────────────────────

@app.get("/frun", response_class=HTMLResponse)
def frun_page(request: Request, msg: str = "", db=Depends(get_db)):
    count   = db.frun_data.count_documents({})
    clients = db.frun_data.distinct("client")
    return tpl(request, "frun.html", {
        "count":   count,
        "clients": sorted(clients),
        "msg":     msg,
    })


@app.post("/frun/upload")
async def upload_frun(request: Request, file: UploadFile = File(...), db=Depends(get_db)):
    ip = _get_client_ip(request)
    content = await file.read()
    try:
        _validate_xlsx(file.filename or "", content, ip)
    except HTTPException as exc:
        security_log.warning(f"FRun upload rejected ip={ip} reason={exc.detail!r}")
        return RedirectResponse(f"/frun?msg={exc.detail}", status_code=303)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        security_log.warning(f"FRun upload parse error ip={ip} filename={file.filename!r} error={e!r}")
        return RedirectResponse(f"/frun?msg=Erreur+fichier:+{e}", status_code=303)

    ws  = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    docs = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        client = str(row[0]).strip() if row[0] else None
        if not client:
            continue
        def _s(v): return str(v).strip() if v is not None else None
        docs.append({
            "client":             client,
            "sid":                _s(row[1]),
            "compliant":          _s(row[2]),
            "landscape":          _s(row[3]),
            "check_description":  _s(row[4]),
            "configuration_item": _s(row[5]),
            "value":              _s(row[6]),
            "policy":             _s(row[7]),
            "check_ref":          clean_check_ref(row[8]),
            "rule":               _s(row[9]),
            "valid_since_utc":    _s(row[10]),
            "imported_at":        now,
            "status":             "Not Started",
            "comment":            "",
            "status_history":     [],
        })

    if docs:
        db.frun_data.insert_many(docs)

    msg = f"{len(docs)} entrées importées le {now}"
    return RedirectResponse(f"/frun?msg={msg}", status_code=303)


# ── Routes : Vue fusionnée ────────────────────────────────────

@app.get("/merged", response_class=HTMLResponse)
def merged_view(
    request: Request,
    client: str = "", sid: str = "", reference: str = "",
    db=Depends(get_db),
):
    client    = _sanitize_text(client, "client")
    sid       = _sanitize_text(sid, "sid")
    reference = _sanitize_text(reference, "reference")
    entries = build_merged_entries(db, client or None, sid or None, reference or None)
    clients = sorted(db.frun_data.distinct("client"))
    return tpl(request, "merged.html", {
        "entries":          entries,
        "status_values":    STATUS_VALUES,
        "filter_client":    client,
        "filter_sid":       sid,
        "filter_reference": reference,
        "clients":          clients,
        "total":            len(entries),
    })


@app.get("/merged/export")
def export_merged(
    client: str = "", sid: str = "", reference: str = "",
    db=Depends(get_db),
):
    from openpyxl.styles import Alignment, Font, PatternFill

    entries = build_merged_entries(db, client or None, sid or None, reference or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VulnTrack"

    COLUMNS = [
        ("Client",              lambda e: e.get("client")),
        ("SID",                 lambda e: e.get("sid")),
        ("Priority",            lambda e: e.get("priority")),
        ("Statut",              lambda e: e.get("status")),
        ("Reference Note",      lambda e: e.get("reference")),
        ("Check Description",   lambda e: e.get("check_description")),
        ("Policy",              lambda e: e.get("policy")),
        ("Correction Type",     lambda e: e.get("correction_type")),
        ("Recommended Process", lambda e: e.get("recommended_implementation_process")),
        ("Downtime Required",   lambda e: e.get("downtime_required")),
        ("CVE",                 lambda e: e.get("cve")),
        ("Solution courte",     lambda e: e.get("solution_short")),
        ("Workaround",          lambda e: e.get("workaround")),
        ("Landscape",           lambda e: e.get("landscape")),
        ("Configuration Item",  lambda e: e.get("configuration_item")),
        ("Value",               lambda e: e.get("value")),
        ("Rule",                lambda e: e.get("rule")),
        ("Valid Since",         lambda e: e.get("valid_since_utc")),
        ("Commentaire",         lambda e: e.get("comment", "")),
    ]
    PRIORITY_COL_IDX = 3

    ws.append([col[0] for col in COLUMNS])

    header_fill = PatternFill("solid", fgColor="FF1E2338")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    for e in entries:
        ws.append([fn(e) for _, fn in COLUMNS])

    PRIO_FILL = {
        "P1": PatternFill("solid", fgColor="FFC0392B"),
        "P2": PatternFill("solid", fgColor="FFD35400"),
        "P3": PatternFill("solid", fgColor="FF27AE60"),
    }
    PRIO_FONT = Font(bold=True, color="FFFFFF")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        prio_cell = row[PRIORITY_COL_IDX - 1]
        fill = PRIO_FILL.get(prio_cell.value)
        if fill:
            prio_cell.fill = fill
            prio_cell.font = PRIO_FONT

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"VulnTrack_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Routes : Entrées (statut, commentaire) ────────────────────

@app.post("/entries/{entry_id}/status")
def update_status(entry_id: str, request_data: dict, db=Depends(get_db)):
    try:
        oid = ObjectId(entry_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID invalide")

    entry = db.frun_data.find_one({"_id": oid})
    if not entry:
        raise HTTPException(status_code=404, detail="Entrée introuvable")

    new_status = request_data.get("status", "")
    if new_status not in STATUS_VALUES:
        raise HTTPException(status_code=400, detail="Statut invalide")

    old_status = entry.get("status", "Not Started")
    db.frun_data.update_one(
        {"_id": oid},
        {
            "$set": {"status": new_status},
            "$push": {"status_history": {
                "from":       old_status,
                "to":         new_status,
                "changed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }},
        },
    )
    return {"ok": True, "status": new_status}


@app.post("/entries/{entry_id}/comment")
def update_comment(entry_id: str, request_data: dict, db=Depends(get_db)):
    try:
        oid = ObjectId(entry_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID invalide")

    comment = _sanitize_text(request_data.get("comment", ""), "comment")
    result  = db.frun_data.update_one({"_id": oid}, {"$set": {"comment": comment}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Entrée introuvable")
    return {"ok": True}


@app.get("/entries/{entry_id}/history")
def get_history(entry_id: str, db=Depends(get_db)):
    try:
        oid = ObjectId(entry_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID invalide")

    entry = db.frun_data.find_one({"_id": oid}, {"status_history": 1})
    if not entry:
        raise HTTPException(status_code=404, detail="Entrée introuvable")
    return {"history": entry.get("status_history", [])}


# ── Routes : Visualisation ────────────────────────────────────

@app.get("/view", response_class=HTMLResponse)
def view(request: Request, client: str = "", db=Depends(get_db)):
    client = _sanitize_text(client, "client")
    all_clients = sorted(db.frun_data.distinct("client"))

    summary = []
    for c in all_clients:
        entries = build_merged_entries(db, client=c)
        summary.append(_client_summary(entries, c))

    all_entries = build_merged_entries(db)
    g_total = len(all_entries)
    g_p1    = sum(1 for e in all_entries if e["priority"] == "P1")
    g_p2    = sum(1 for e in all_entries if e["priority"] == "P2")
    g_p3    = sum(1 for e in all_entries if e["priority"] == "P3")
    g_done  = sum(1 for e in all_entries if e["status"] in ("Validated", "Implemented"))
    g_ns    = sum(1 for e in all_entries if e["status"] == "Not Started")

    pipeline_order = ["Not Started", "Started", "Implemented",
                      "Validated", "To be checked", "Exception"]
    g_by_status: dict[str, int] = {}
    for e in all_entries:
        g_by_status[e["status"]] = g_by_status.get(e["status"], 0) + 1

    global_stats = {
        "total": g_total, "p1": g_p1, "p2": g_p2, "p3": g_p3,
        "done": g_done, "not_started": g_ns,
        "progress_pct": round(g_done / g_total * 100) if g_total else 0,
        "pipeline": [{"label": s, "count": g_by_status.get(s, 0)} for s in pipeline_order],
        "clients": len(all_clients),
    }

    selected_entries = []
    selected_summary = None
    if client:
        selected_entries = build_merged_entries(db, client=client)
        selected_summary = next((s for s in summary if s["client"] == client), None)

    return tpl(request, "view.html", {
        "summary":          summary,
        "global_stats":     global_stats,
        "selected_client":  client,
        "selected_entries": selected_entries,
        "selected_summary": selected_summary,
        "status_values":    STATUS_VALUES,
        "all_clients":      all_clients,
        "pipeline_order":   pipeline_order,
    })
